"""
CSV/Excel processor using Polars.
Cleans data and generates a two-sheet Excel report (Summary + Clean Data).
"""

import json
from datetime import datetime
from pathlib import Path

import polars as pl


def load_file(path: Path) -> pl.DataFrame:
    """Load a CSV or Excel file into a Polars DataFrame."""
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return pl.read_csv(path, infer_schema_length=500, ignore_errors=True)
    elif suffix in (".xlsx", ".xls"):
        return pl.read_excel(path)
    else:
        raise ValueError(f"Unsupported file type: {suffix}. Accepted: .csv, .xlsx, .xls")


def clean(df: pl.DataFrame) -> tuple[pl.DataFrame, dict]:
    """
    Clean a DataFrame and return (clean_df, stats).

    Steps:
    1. Strip leading/trailing whitespace from string columns
    2. Remove fully duplicate rows
    3. Fill nulls: numeric → 0, string → "N/A"
    """
    rows_raw = len(df)

    # 1. Strip whitespace from string columns
    str_cols = [c for c, t in zip(df.columns, df.dtypes) if t in (pl.Utf8, pl.String)]
    if str_cols:
        df = df.with_columns([pl.col(c).str.strip_chars() for c in str_cols])

    # 2. Remove duplicates
    df_dedup = df.unique()
    duplicates_removed = rows_raw - len(df_dedup)
    df = df_dedup

    # 3. Count nulls before fill
    null_counts = {c: df[c].null_count() for c in df.columns}
    total_nulls = sum(null_counts.values())

    # 4. Fill nulls
    fill_exprs = []
    for col_name, dtype in zip(df.columns, df.dtypes):
        if dtype in (pl.Int32, pl.Int64, pl.Float32, pl.Float64):
            fill_exprs.append(pl.col(col_name).fill_null(0))
        elif dtype in (pl.Utf8, pl.String):
            fill_exprs.append(pl.col(col_name).fill_null("N/A"))
    if fill_exprs:
        df = df.with_columns(fill_exprs)

    rows_clean = len(df)

    stats = {
        "rows_raw": rows_raw,
        "rows_clean": rows_clean,
        "duplicates_removed": duplicates_removed,
        "nulls_filled": total_nulls,
        "columns": df.columns,
        "null_counts_before": null_counts,
        "column_count": len(df.columns),
    }

    return df, stats


def generate_report(df: pl.DataFrame, stats: dict, job_id: int, reports_dir: Path) -> Path:
    """
    Generate a two-sheet Excel report:
    - Sheet 1 "Summary": processing stats overview
    - Sheet 2 "Clean Data": the cleaned rows

    Also saves a JSON sidecar for programmatic access.
    """
    reports_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    excel_path = reports_dir / f"report_{job_id}_{timestamp}.xlsx"

    # Write clean data first
    df.write_excel(excel_path, worksheet="Clean Data")

    # Add Summary sheet via openpyxl
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.load_workbook(excel_path)

    # Create Summary sheet at position 0
    ws_summary = wb.create_sheet("Summary", 0)

    # Header style
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor="1E3A5F")
    value_fill   = PatternFill("solid", fgColor="F0F4F8")

    rows = [
        ("Field",                  "Value"),
        ("Job ID",                 job_id),
        ("Processed At (UTC)",     timestamp.replace("_", " ")),
        ("Original Rows",          stats["rows_raw"]),
        ("Clean Rows",             stats["rows_clean"]),
        ("Duplicates Removed",     stats["duplicates_removed"]),
        ("Missing Values Fixed",   stats["nulls_filled"]),
        ("Columns Processed",      stats.get("column_count", len(stats.get("columns", [])))),
        ("Column Names",           ", ".join(stats["columns"])),
        ("Data Quality Improvement",
         f"{(stats['duplicates_removed'] + stats['nulls_filled'])} issues resolved"),
    ]

    for i, (field, value) in enumerate(rows, start=1):
        cell_a = ws_summary.cell(row=i, column=1, value=field)
        cell_b = ws_summary.cell(row=i, column=2, value=value)
        if i == 1:
            cell_a.font = header_font
            cell_b.font = header_font
            cell_a.fill = header_fill
            cell_b.fill = header_fill
        else:
            cell_b.fill = value_fill
        cell_a.alignment = Alignment(vertical="center")
        cell_b.alignment = Alignment(vertical="center")

    ws_summary.column_dimensions["A"].width = 28
    ws_summary.column_dimensions["B"].width = 40

    wb.save(excel_path)

    # JSON sidecar
    report_data = {
        "job_id": job_id,
        "generated_at": timestamp,
        "rows_raw": stats["rows_raw"],
        "rows_clean": stats["rows_clean"],
        "duplicates_removed": stats["duplicates_removed"],
        "nulls_filled": stats["nulls_filled"],
        "columns": stats["columns"],
    }
    json_path = reports_dir / f"report_{job_id}_{timestamp}.json"
    json_path.write_text(json.dumps(report_data, indent=2, ensure_ascii=False))

    return excel_path


def process_file(file_path: Path, job_id: int, reports_dir: Path) -> dict:
    """Full pipeline: load → clean → report. Returns stats dict with report_path."""
    df_raw  = load_file(file_path)
    df_clean, stats = clean(df_raw)
    report_path = generate_report(df_clean, stats, job_id, reports_dir)
    stats["report_path"] = str(report_path)
    return stats
